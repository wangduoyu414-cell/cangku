from __future__ import annotations

from pathlib import Path
import argparse
import json
import subprocess
import sys
import tempfile
from collections import Counter

from openpyxl import load_workbook
import yaml


DEFAULT_SKILL_ROOT = Path(r"D:\agent\skill\language-adapter-dataset-builder")


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def load_yaml(path: Path) -> dict:
    return yaml.safe_load(path.read_text(encoding="utf-8"))


def dump_yaml(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(yaml.safe_dump(payload, allow_unicode=True, sort_keys=False), encoding="utf-8")


def parse_workbook(path: Path) -> tuple[dict, dict]:
    workbook = load_workbook(path)
    package_ws = workbook.worksheets[1]
    items_ws = workbook.worksheets[2]

    package = {}
    for row in range(2, 7):
        key = str(package_ws.cell(row, 1).value or "")
        value = package_ws.cell(row, 3).value
        if "pack_id" in key:
            package["pack_id"] = value
        elif "source_model" in key:
            package["source_model"] = value
        elif "source_ref" in key:
            package["source_ref"] = value
        elif "generated_at" in key:
            package["generated_at"] = value
        elif "rule_version_id" in key:
            package["rule_version_id"] = value

    rows = []
    self_check_counter: Counter[str] = Counter()
    label_rule_pass_rows: list[int] = []
    for row in range(2, items_ws.max_row + 1):
        values = [items_ws.cell(row, column).value for column in range(1, 31)]
        if not any(value not in (None, "") for value in values):
            continue
        task_type = values[2]
        candidate_type = values[1]
        trace_refs = [item for item in [values[4], values[5]] if item not in (None, "")]
        content = {
            "sample_id": values[6],
            "case_id": values[7],
        }
        if task_type == "comparison_explanation":
            content["input"] = {
                "primary_candidate_id": values[8],
                "fallback_candidate_id": values[9],
                "lead_summary": values[10],
                "primary_reasons": json.loads(values[11]) if values[11] else [],
                "fallback_tradeoffs": json.loads(values[12]) if values[12] else [],
                "primary_risks": json.loads(values[13]) if values[13] else [],
                "counterfactuals": json.loads(values[14]) if values[14] else [],
            }
            content["target"] = {"target_text": values[15]}
        elif task_type == "safety_refusal":
            content["input"] = {
                "raw_input": values[16],
                "reason_codes": json.loads(values[17]) if values[17] else [],
                "redirect": values[18],
                "risk_level": values[19],
            }
            content["target"] = {"refusal_text": values[20]}
        elif task_type == "supplemental_parse":
            content["input"] = {
                "input_text": values[21],
                "normalized_fields": json.loads(values[22]) if values[22] else {},
                "conflict_candidates": json.loads(values[23]) if values[23] else [],
                "confidence_notes": values[24],
            }
            content["target"] = {
                "structured_delta": json.loads(values[25]) if values[25] else {},
                "anchors": json.loads(values[26]) if values[26] else [],
                "abstain": True if values[27] == "true" else False if values[27] == "false" else None,
                "source_kind": "placeholder",
                "evidence_refs": list(trace_refs),
                "source_artifact": values[3],
                "fidelity_checks": {},
                "open_risks": [],
            }

        rows.append(
            {
                "row": row,
                "candidate_id": values[0],
                "candidate_type": candidate_type,
                "task_type": task_type,
                "source_ref": values[3],
                "trace_refs": trace_refs,
                "content_notes": values[28],
                "self_check": values[29],
                "content": content,
            }
        )
        self_check_counter.update([str(values[29] or "")])
        if candidate_type == "label_rule" and str(values[29] or "").strip() == "pass":
            label_rule_pass_rows.append(row)

    return (
        {"candidate_pack": {**package, "items": [{key: value for key, value in row.items() if key not in {"row", "content_notes", "self_check"}} for row in rows]}},
        {
            "rows": rows,
            "self_check_counter": dict(self_check_counter),
            "label_rule_pass_rows": label_rule_pass_rows,
        },
    )


def parse_candidate_input(path: Path) -> tuple[dict, dict]:
    if path.suffix.lower() == ".xlsx":
        return parse_workbook(path)
    payload = load_json(path)
    rows = []
    for index, item in enumerate(payload.get("candidate_pack", {}).get("items", []), start=1):
        rows.append(
            {
                "row": index,
                "candidate_id": item.get("candidate_id", ""),
                "candidate_type": item.get("candidate_type", ""),
                "task_type": item.get("task_type", ""),
                "source_ref": item.get("source_ref", ""),
                "trace_refs": item.get("trace_refs", []),
                "content_notes": "",
                "self_check": "",
                "content": item.get("content", {}),
            }
        )
    return payload, {"rows": rows, "self_check_counter": {}, "label_rule_pass_rows": []}


def run_validator(candidate_pack_payload: dict, skill_root: Path) -> dict:
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_root = Path(temp_dir)
        candidate_path = temp_root / "candidate_pack.json"
        output_path = temp_root / "validated_supplement.json"
        candidate_path.write_text(json.dumps(candidate_pack_payload, ensure_ascii=False, indent=2), encoding="utf-8")
        subprocess.run(
            [
                sys.executable,
                str(skill_root / "scripts" / "validate_candidate_pack.py"),
                "--candidate-pack",
                str(candidate_path),
                "--output",
                str(output_path),
            ],
            check=True,
        )
        return load_json(output_path)


def compare_manifest(manifest_payload: dict | None, candidate_pack_payload: dict) -> dict:
    if manifest_payload is None:
        return {"matched": False, "reason": "未提供 run manifest（运行清单），跳过配额对比。"}

    manifest = manifest_payload.get("run_manifest", {})
    pack_plan = manifest.get("pack_plan", [])
    pack_id = candidate_pack_payload.get("candidate_pack", {}).get("pack_id", "")
    pack_item = next((item for item in pack_plan if item.get("pack_id") == pack_id), None)
    if pack_item is None:
        return {"matched": False, "reason": f"run manifest（运行清单） 中未找到 pack_id={pack_id!r}。"}

    actual_counter = Counter(
        (
            item.get("task_type", ""),
            item.get("candidate_type", ""),
        )
        for item in candidate_pack_payload.get("candidate_pack", {}).get("items", [])
    )
    expected_counter = Counter(
        {
            (item.get("task_type", ""), item.get("candidate_type", "")): int(item.get("count", 0))
            for item in pack_item.get("expected_mix", [])
        }
    )

    mismatches = []
    for key in sorted(set(actual_counter) | set(expected_counter)):
        if actual_counter.get(key, 0) != expected_counter.get(key, 0):
            mismatches.append(
                {
                    "task_type": key[0],
                    "candidate_type": key[1],
                    "expected": expected_counter.get(key, 0),
                    "actual": actual_counter.get(key, 0),
                }
            )
    return {
        "matched": len(mismatches) == 0,
        "pack_id": pack_id,
        "mismatches": mismatches,
        "required_cases": pack_item.get("required_cases", []),
        "required_personas": pack_item.get("required_personas", []),
        "required_need_axes": pack_item.get("required_need_axes", []),
        "required_difficulty_axes": pack_item.get("required_difficulty_axes", []),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="evaluate candidate batch（测评候选批次）")
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--manifest", type=Path)
    parser.add_argument("--skill-root", type=Path, default=DEFAULT_SKILL_ROOT)
    args = parser.parse_args()

    candidate_pack_payload, workbook_meta = parse_candidate_input(args.input)
    validator_payload = run_validator(candidate_pack_payload, args.skill_root)
    manifest_payload = load_yaml(args.manifest) if args.manifest else None
    manifest_check = compare_manifest(manifest_payload, candidate_pack_payload)

    rows = workbook_meta["rows"]
    pair_counter = Counter((item["task_type"], item["candidate_type"]) for item in rows)
    task_counter = Counter(item["task_type"] for item in rows)
    validator_items = validator_payload.get("validated_supplement", {}).get("items", [])
    decision_counter = Counter(item.get("decision", "") for item in validator_items)
    scope_counter = Counter(item.get("adoption_scope", "") for item in validator_items)

    source_ref = str(candidate_pack_payload.get("candidate_pack", {}).get("source_ref", "")).strip()
    demo_source_detected = source_ref.startswith("candidate://demo")
    ungrounded_trace_rows = []
    for item in rows:
        refs = [str(ref).strip() for ref in item["trace_refs"] if str(ref).strip()]
        has_replay_key = any(ref.startswith("replay://") for ref in refs)
        has_real_path = any(Path(ref).exists() for ref in refs if ":" in ref or ref.startswith("\\"))
        if not (has_replay_key and has_real_path):
            ungrounded_trace_rows.append(item["row"])
    validator_rejected = [item.get("candidate_id", "") for item in validator_items if item.get("decision") == "rejected"]
    validator_needs_review = [item.get("candidate_id", "") for item in validator_items if item.get("decision") == "needs_review"]

    blockers = []
    warnings = []
    if demo_source_detected:
        blockers.append(
            {
                "code": "demo_source_detected",
                "message": "当前批次仍使用 demo（演示） 来源，不能作为正式候选批次。",
                "evidence_refs": [str(args.input)],
            }
        )
    if ungrounded_trace_rows:
        blockers.append(
            {
                "code": "ungrounded_trace_detected",
                "message": f"存在 {len(ungrounded_trace_rows)} 行追踪引用未落到真实 replay（回放） 或 artifact（产物） 语义。",
                "evidence_refs": [str(args.input)],
            }
        )
    if workbook_meta["label_rule_pass_rows"]:
        blockers.append(
            {
                "code": "label_rule_self_check_mismatch",
                "message": f"label_rule（标注规则） 行被错误标记为 pass（通过），行号={workbook_meta['label_rule_pass_rows']}",
                "evidence_refs": [str(args.input)],
            }
        )
    if not manifest_check.get("matched", True):
        if manifest_check.get("mismatches"):
            blockers.append(
                {
                    "code": "mix_quota_mismatch",
                    "message": "候选配额与运行清单不一致。",
                    "evidence_refs": [str(args.manifest)] if args.manifest else [str(args.input)],
                }
            )
        elif manifest_check.get("reason"):
            warnings.append(manifest_check["reason"])
    if validator_rejected:
        blockers.append(
            {
                "code": "validator_rejected_items",
                "message": f"下层 validate_candidate_pack（候选包校验） 出现 rejected（拒绝） 条目：{validator_rejected[:10]}",
                "evidence_refs": [str(args.input)],
            }
        )
    if validator_needs_review:
        blockers.append(
            {
                "code": "validator_needs_review_items",
                "message": f"下层 validate_candidate_pack（候选包校验） 出现 needs_review（待审） 条目：{validator_needs_review[:10]}",
                "evidence_refs": [str(args.input)],
            }
        )

    status = "fail" if blockers else ("warn" if warnings else "pass")
    report = {
        "evaluation_report": {
            "agent_id": "language-adapter-orchestration-evaluator",
            "status": status,
            "input_ref": str(args.input),
            "manifest_ref": str(args.manifest) if args.manifest else "",
            "summary": {
                "total_items": len(rows),
                "task_counts": {key: value for key, value in task_counter.items()},
                "pair_counts": {f"{task_type}/{candidate_type}": count for (task_type, candidate_type), count in pair_counter.items()},
                "validator_decisions": {key: value for key, value in decision_counter.items()},
                "validator_scopes": {key: value for key, value in scope_counter.items()},
                "self_check_counts": workbook_meta["self_check_counter"],
            },
            "manifest_check": manifest_check,
            "blockers": blockers,
            "warnings": warnings,
            "evidence_refs": [str(args.input)] + ([str(args.manifest)] if args.manifest else []),
        }
    }
    dump_yaml(args.output, report)
    print(f"evaluation report（测评报告）: {args.output}")
    return 1 if status == "fail" else 0


if __name__ == "__main__":
    raise SystemExit(main())
